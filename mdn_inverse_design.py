import os, time
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import TensorDataset, DataLoader
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler, MinMaxScaler

# ========== 설정 ==========
DATA_PATH   = r'D:\이영재\Inverse design\lab_result.xlsx'
INPUT_TYPE  = 'lab'     
NUM_PARAMS  = 5          
NUM_SPECTRUM= 301        
LAB_COLS    = ['L','a','b'] 
K_MIXTURES  = 15        
BATCH_SIZE  = 16
EPOCHS      = 2000
LR          = 1e-3
VAL_SPLIT   = 0.15
TEST_SPLIT  = 0.15
SEED        = 42
SAVE_PREFIX = 'mdn'     

torch.manual_seed(SEED)
np.random.seed(SEED)

# ========== 데이터 로드 ==========
df = pd.read_excel(DATA_PATH)

Y = df.iloc[:, :NUM_PARAMS].values.astype(np.float32)

if INPUT_TYPE.lower() == 'lab':
    assert all([c in df.columns for c in LAB_COLS]), f"Lab 컬럼 {LAB_COLS} 가 데이터에 없습니다."
    X = df[LAB_COLS].values.astype(np.float32)  
    IN_DIM = X.shape[1]  
elif INPUT_TYPE.lower() == 'spectrum':
    X = df.iloc[:, NUM_PARAMS:NUM_PARAMS+NUM_SPECTRUM].values.astype(np.float32)
    IN_DIM = X.shape[1]
else:
    raise ValueError("INPUT_TYPE 은 'lab' 또는 'spectrum' 이어야 합니다.")


x_scaler = StandardScaler()
X_scaled = x_scaler.fit_transform(X).astype(np.float32)

y_scaler = MinMaxScaler()
Y_scaled = y_scaler.fit_transform(Y).astype(np.float32)

X_trainval, X_test, Y_trainval, Y_test = train_test_split(
    X_scaled, Y_scaled, test_size=TEST_SPLIT, random_state=SEED
)
val_ratio = VAL_SPLIT / (1.0 - TEST_SPLIT)
X_train, X_val, Y_train, Y_val = train_test_split(
    X_trainval, Y_trainval, test_size=val_ratio, random_state=SEED
)

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
X_train_t = torch.tensor(X_train, dtype=torch.float32)
Y_train_t = torch.tensor(Y_train, dtype=torch.float32)
X_val_t   = torch.tensor(X_val,   dtype=torch.float32)
Y_val_t   = torch.tensor(Y_val,   dtype=torch.float32)
X_test_t  = torch.tensor(X_test,  dtype=torch.float32)
Y_test_t  = torch.tensor(Y_test,  dtype=torch.float32)

train_loader = DataLoader(TensorDataset(X_train_t, Y_train_t), batch_size=BATCH_SIZE, shuffle=True)
val_loader   = DataLoader(TensorDataset(X_val_t,   Y_val_t),   batch_size=BATCH_SIZE, shuffle=False)

# ========== 모델 ==========
class MDNHead(nn.Module):
    def __init__(self, hidden_dim, num_params=5, K=16, use_mu_sigmoid=True):
        super().__init__()
        self.P = num_params
        self.K = K
        self.use_mu_sigmoid = use_mu_sigmoid
        self.out = nn.Linear(hidden_dim, num_params * K * 3)
        self.softplus = nn.Softplus()

    def forward(self, h):
        B = h.size(0)
        raw = self.out(h)                     
        raw = raw.view(B, self.P, self.K, 3)    
        mu_raw     = raw[..., 0]                 
        sigma_raw  = raw[..., 1]
        pi_logits  = raw[..., 2]
        mu = torch.sigmoid(mu_raw) if self.use_mu_sigmoid else mu_raw
        sigma = self.softplus(sigma_raw) + 1e-6
        pi = F.softmax(pi_logits, dim=-1)
        return mu, sigma, pi

def mdn_nll(y_true, mu, sigma, pi):
    B,P,K = mu.shape
    y = y_true.unsqueeze(-1).expand(-1, -1, K)              
    log_norm = -0.5*np.log(2*np.pi) - torch.log(sigma)       
    log_exp  = -0.5*((y - mu)/sigma)**2                       
    log_prob = log_norm + log_exp                            
    log_mix  = torch.log(pi + 1e-12) + log_prob
    log_sum  = torch.logsumexp(log_mix, dim=-1)   
    nll = -(log_sum).mean()
    return nll

def mdn_map_decode(mu, sigma, pi):
    with torch.no_grad():
        k_star = pi.argmax(dim=-1, keepdim=True)       
        mu_star = torch.gather(mu, dim=-1, index=k_star).squeeze(-1)
    return mu_star

def mdn_sample(mu, sigma, pi, n_samples=50):
    B,P,K = mu.shape
    out = []
    with torch.no_grad():
        cat = torch.distributions.Categorical(probs=pi)
        for _ in range(n_samples):
            k = cat.sample()                             
            idx = k.unsqueeze(-1)                        
            mu_sel    = torch.gather(mu,    -1, idx).squeeze(-1)
            sigma_sel = torch.gather(sigma, -1, idx).squeeze(-1)  
            eps = torch.randn_like(mu_sel)
            y_samp = mu_sel + sigma_sel * eps
            y_samp = torch.clamp(y_samp, 0.0, 1.0)
            out.append(y_samp)
    return torch.stack(out, dim=0)                

class BackboneMLP(nn.Module):
    def __init__(self, in_dim, hidden=128):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(in_dim, 64), nn.ReLU(),
            nn.Linear(64, 128),   nn.ReLU(),
            nn.Linear(128, hidden), nn.ReLU(),
        )
    def forward(self, x):
        return self.net(x)

class MDNModel(nn.Module):
    def __init__(self, in_dim, num_params=5, K=16, hidden=128, use_mu_sigmoid=True):
        super().__init__()
        self.backbone = BackboneMLP(in_dim, hidden=hidden)
        self.mdn = MDNHead(hidden, num_params=num_params, K=K, use_mu_sigmoid=use_mu_sigmoid)
    def forward(self, x):
        h = self.backbone(x)
        return self.mdn(h)

model = MDNModel(IN_DIM, num_params=NUM_PARAMS, K=K_MIXTURES, hidden=128, use_mu_sigmoid=True).to(device)

# ========== 학습 ==========
optimizer = torch.optim.Adam(model.parameters(), lr=LR)
scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='min', factor=0.5, patience=20)

history = {'epoch':[], 'train_nll':[], 'val_nll':[], 'val_map_rmse':[]}

def rmse(a, b):
    return np.sqrt(np.mean((a-b)**2))

best_val = float('inf')
start = time.time()
for epoch in range(1, EPOCHS+1):
    model.train()
    nll_accum = 0.0
    n_batch = 0
    for xb, yb in train_loader:
        xb, yb = xb.to(device), yb.to(device)
        mu, sigma, pi = model(xb)
        loss = mdn_nll(yb, mu, sigma, pi)
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        nll_accum += loss.item()
        n_batch += 1
    train_nll = nll_accum / max(1, n_batch)

    # validation
    model.eval()
    with torch.no_grad():
        val_nll_accum, nb = 0.0, 0
        preds_map, truths = [], []
        for xb, yb in val_loader:
            xb, yb = xb.to(device), yb.to(device)
            mu, sigma, pi = model(xb)
            loss = mdn_nll(yb, mu, sigma, pi)
            val_nll_accum += loss.item()
            nb += 1
            y_map = mdn_map_decode(mu, sigma, pi)           
            y_map_np = y_map.cpu().numpy()
            y_map_denorm = y_scaler.inverse_transform(y_map_np)
            y_true_denorm = y_scaler.inverse_transform(yb.cpu().numpy())
            preds_map.append(y_map_denorm)
            truths.append(y_true_denorm)
        val_nll = val_nll_accum / max(1, nb)
        preds_map = np.vstack(preds_map)
        truths    = np.vstack(truths)
        val_map_rmse = rmse(preds_map, truths)

    scheduler.step(val_nll)

    history['epoch'].append(epoch)
    history['train_nll'].append(train_nll)
    history['val_nll'].append(val_nll)
    history['val_map_rmse'].append(val_map_rmse)

    if epoch % 50 == 0 or epoch == 1:
        print(f"[{epoch:04d}] train NLL={train_nll:.4f} | val NLL={val_nll:.4f} | val MAP RMSE={val_map_rmse:.4f}")

    if val_nll < best_val:
        best_val = val_nll
        torch.save({'model': model.state_dict(),
                    'x_scaler_mean': x_scaler.mean_,
                    'x_scaler_scale': x_scaler.scale_,
                    'y_scaler_min': y_scaler.min_,
                    'y_scaler_scale': y_scaler.scale_}, f'{SAVE_PREFIX}_best.pt')

elapsed = time.time() - start
print(f"학습 완료: {elapsed/60:.1f}분, best val NLL={best_val:.4f}")

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

# 1. Test
model.eval()
with torch.no_grad():
    mu, sigma, pi = model(X_test_t.to(device))
    y_map = mdn_map_decode(mu, sigma, pi).cpu().numpy()   
    y_pred = y_scaler.inverse_transform(y_map)             
    y_true = y_scaler.inverse_transform(Y_test_t.cpu().numpy())  
    test_rmse_overall = rmse(y_pred, y_true)
    per_param_rmse = np.sqrt(np.mean((y_pred - y_true)**2, axis=0))

print(f"Test RMSE (MAP, overall): {test_rmse_overall:.6f}")
print("Per-Param RMSE:", per_param_rmse)

# 2. Forward FNN 불러오기
from pathlib import Path as _Path 

ckpt_path = r"D:\이영재\Inverse design\FNN_Lab.pth"  

def build_forward_fnn(in_dim=5, out_dim=3):
    return nn.Sequential(
        nn.Linear(in_dim, 64),
        nn.ReLU(),
        nn.Linear(64, 128),
        nn.ReLU(),
        nn.Linear(128, 128),
        nn.ReLU(),
        nn.Linear(128, 128),
        nn.ReLU(),
        nn.Linear(128, out_dim)
    )
ckpt = torch.load(ckpt_path, map_location=device, weights_only=False)
state_dict_fwd = ckpt['model_state']
scaler_X_fwd   = ckpt['scaler_X']   
scaler_Y_fwd   = ckpt['scaler_Y']   

forward_model = build_forward_fnn(in_dim=NUM_PARAMS, out_dim=len(LAB_COLS)).to(device)
forward_model.load_state_dict(state_dict_fwd)
forward_model.eval()

# 3. Forward FNN으로 Lab 예측
X_forward_scaled = scaler_X_fwd.transform(y_pred.astype(np.float32)) 
with torch.no_grad():
    X_forward_tensor = torch.tensor(X_forward_scaled, dtype=torch.float32, device=device)
    Lab_pred_scaled = forward_model(X_forward_tensor).cpu().numpy() 
Lab_pred = scaler_Y_fwd.inverse_transform(Lab_pred_scaled)        

# 4. 실제 Lab 
if INPUT_TYPE.lower() == 'lab':
    Lab_true = x_scaler.inverse_transform(X_test_t.cpu().numpy())
else:
    Lab_true = None

# 5. delta E 계산
def deltaE(Lab1, Lab2):
    diff = Lab1 - Lab2
    return np.sqrt(np.sum(diff**2, axis=-1))

if Lab_true is not None:
    deltaE_vals = deltaE(Lab_true, Lab_pred)
else:
    deltaE_vals = np.full((y_pred.shape[0],), np.nan, dtype=np.float32)

# 6. 결과 출력
param_names = ['r_pore', 'p_period', 't_AAO', 't_Ag_top', 't_Ag_bot']

df_actual = pd.DataFrame(y_true, columns=[f'Actual_{n}' for n in param_names])
df_pred   = pd.DataFrame(y_pred, columns=[f'PredMAP_{n}' for n in param_names])

if Lab_true is not None:
    df_lab_true  = pd.DataFrame(Lab_true,  columns=[f'Lab_true_{c}' for c in LAB_COLS])
    df_lab_pred  = pd.DataFrame(Lab_pred,  columns=[f'Lab_pred_{c}' for c in LAB_COLS])
    df_deltaE    = pd.DataFrame(deltaE_vals, columns=['DeltaE'])
    df_all = pd.concat([df_lab_true, df_lab_pred, df_deltaE, df_actual, df_pred], axis=1)
else:
    df_all = pd.concat([df_actual, df_pred], axis=1)

if Lab_true is not None:
    good_mask = deltaE_vals < 1.0
    good_rows = df_all[good_mask]
    print("ΔE < 1.0 인 샘플들:")
    print(good_rows)

# 7. 엑셀 저장
export_path = f'{SAVE_PREFIX}_results.xlsx'
df_all.to_excel(export_path, index=False)

wb = openpyxl.load_workbook(export_path)
ws = wb.active

deltaE_col_idx = None
for idx, cell in enumerate(ws[1], start=1):
    if cell.value == 'DeltaE':
        deltaE_col_idx = idx
        break

if deltaE_col_idx is not None:
    highlight_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=deltaE_col_idx)
        val = cell.value
        try:
            if float(val) < 1.0:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = highlight_fill
        except:
            pass

wb.save(export_path)
print(f"엑셀 저장 완료: {export_path}")
